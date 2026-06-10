import os
import json
from pathlib import Path
from typing import List, Tuple, Optional

from pypdf import PdfReader, PdfWriter
import pikepdf

from core.tempfiles import generate_temp_path, cleanup_temp, save_upload
from core.utils import parse_page_range
from core.errors import raise_bad_request, raise_internal_error


def compress_pdf_service(input_path: Path, output_path: Path) -> Tuple[Path, float]:
    with pikepdf.Pdf.open(input_path) as pdf:
        pdf.save(output_path, compress_streams=True,
                 object_stream_mode=pikepdf.ObjectStreamMode.generate)

    original_size = os.path.getsize(input_path)
    compressed_size = os.path.getsize(output_path)

    if compressed_size >= original_size:
        os.remove(output_path)
        return input_path, 0.0

    reduction = ((original_size - compressed_size) / original_size) * 100
    os.remove(input_path)
    return output_path, reduction


def _oriented_page(page, orientation: str):
    if orientation not in ("portrait", "landscape"):
        return page

    box = page.mediabox
    width = float(box.width)
    height = float(box.height)
    if orientation == "portrait" and width > height:
        page.rotate(90)
    elif orientation == "landscape" and height > width:
        page.rotate(90)
    return page


def merge_pdf_service(pdf_paths: List[Path], output_path: Path, orientations: List[str] | None = None, rotations: List[int] | None = None) -> int:
    merger = PdfWriter()
    for index, pdf_path in enumerate(pdf_paths):
        orientation = orientations[index] if orientations and index < len(orientations) else "auto"
        rotation = rotations[index] if rotations and index < len(rotations) else 0
        reader = PdfReader(pdf_path)
        for page in reader.pages:
            page = _oriented_page(page, orientation)
            if rotation:
                page.rotate(rotation)
            merger.add_page(page)

    with open(output_path, "wb") as f:
        merger.write(f)

    total_pages = len(merger.pages)
    cleanup_temp(*pdf_paths)
    return total_pages


def split_pdf_service(
    input_path: Path,
    output_folder: Path,
    mode: str,
    ranges: str
) -> Tuple[List[Path], int]:
    reader = PdfReader(input_path)
    total_pages = len(reader.pages)
    output_files = []

    if mode == "all":
        for i in range(total_pages):
            writer = PdfWriter()
            writer.add_page(reader.pages[i])
            out = output_folder / f"page_{i + 1}.pdf"
            with open(out, "wb") as f:
                writer.write(f)
            output_files.append(out)
    elif mode == "range":
        for range_str in ranges.split(','):
            writer = PdfWriter()
            range_str = range_str.strip()
            if '-' in range_str:
                start, end = map(int, range_str.split('-'))
                for pn in range(start - 1, min(end, total_pages)):
                    writer.add_page(reader.pages[pn])
                out = output_folder / f"pages_{start}-{end}.pdf"
            else:
                pn = int(range_str) - 1
                if 0 <= pn < total_pages:
                    writer.add_page(reader.pages[pn])
                out = output_folder / f"page_{range_str}.pdf"
            with open(out, "wb") as f:
                writer.write(f)
            output_files.append(out)
    elif mode == "specific":
        for pn in [int(p.strip()) for p in ranges.split(',')]:
            if 1 <= pn <= total_pages:
                writer = PdfWriter()
                writer.add_page(reader.pages[pn - 1])
                out = output_folder / f"page_{pn}.pdf"
                with open(out, "wb") as f:
                    writer.write(f)
                output_files.append(out)

    return output_files, total_pages


def reorder_pdf_service(input_path: Path, output_path: Path, order_str: str) -> int:
    try:
        new_order = json.loads(order_str)
    except json.JSONDecodeError:
        raise_bad_request("Formato de ordem invalido. Use JSON array.")

    reader = PdfReader(input_path)
    total_pages = len(reader.pages)

    if len(new_order) != total_pages:
        raise_bad_request(f"A ordem deve conter exatamente {total_pages} paginas")

    if sorted(new_order) != list(range(1, total_pages + 1)):
        raise_bad_request("A ordem deve conter todas as paginas de 1 a N, sem duplicatas ou faltas")

    for p in new_order:
        if p < 1 or p > total_pages:
            raise_bad_request(f"Pagina {p} fora do intervalo (1 a {total_pages})")

    writer = PdfWriter()
    for page_num in new_order:
        writer.add_page(reader.pages[page_num - 1])

    with open(output_path, "wb") as f:
        writer.write(f)

    cleanup_temp(input_path)
    return total_pages


def rotate_pdf_service(input_path: Path, output_path: Path, rotation: int, pages: str) -> Tuple[int, int]:
    if rotation not in (90, 180, 270):
        raise_bad_request("Rotacao invalida. Use 90, 180 ou 270")

    reader = PdfReader(input_path)
    total_pages = len(reader.pages)
    page_list = parse_page_range(pages, total_pages)

    writer = PdfWriter()
    for i in range(total_pages):
        page = reader.pages[i]
        if (i + 1) in page_list:
            page.rotate(rotation)
        writer.add_page(page)

    with open(output_path, "wb") as f:
        writer.write(f)

    cleanup_temp(input_path)
    return total_pages, len(page_list)


def extract_pages_service(input_path: Path, output_path: Path, pages: str) -> Tuple[int, int]:
    reader = PdfReader(input_path)
    total_pages = len(reader.pages)
    page_list = parse_page_range(pages, total_pages)

    if not page_list:
        raise_bad_request("Nenhuma pagina valida informada")

    writer = PdfWriter()
    for pn in page_list:
        writer.add_page(reader.pages[pn - 1])

    with open(output_path, "wb") as f:
        writer.write(f)

    cleanup_temp(input_path)
    return total_pages, len(page_list)


def remove_pages_service(input_path: Path, output_path: Path, pages_to_remove: str) -> Tuple[int, int]:
    reader = PdfReader(input_path)
    total_pages = len(reader.pages)
    remove_list = parse_page_range(pages_to_remove, total_pages)

    if len(remove_list) >= total_pages:
        raise_bad_request("Nao e possivel remover todas as paginas")

    remove_set = set(remove_list)
    writer = PdfWriter()
    kept = 0
    for i in range(total_pages):
        if (i + 1) not in remove_set:
            writer.add_page(reader.pages[i])
            kept += 1

    if kept == 0:
        raise_bad_request("Nao e possivel remover todas as paginas")

    with open(output_path, "wb") as f:
        writer.write(f)

    cleanup_temp(input_path)
    return total_pages, len(remove_list)


def protect_pdf_service(input_path: Path, output_path: Path, password: str) -> Tuple[int, int]:
    if len(password) < 4:
        raise_bad_request("A senha deve ter pelo menos 4 caracteres")

    original_size = os.path.getsize(input_path)
    pdf = pikepdf.Pdf.open(input_path)
    pdf.save(output_path, encryption=pikepdf.Encryption(user=password, owner=password))
    pdf.close()

    protected_size = os.path.getsize(output_path)
    cleanup_temp(input_path)
    return original_size, protected_size


def unlock_pdf_service(input_path: Path, output_path: Path, password: str) -> Tuple[int, int]:
    original_size = os.path.getsize(input_path)

    try:
        pdf = pikepdf.Pdf.open(input_path, password=password)
    except pikepdf.PasswordError:
        from core.errors import raise_bad_password
        raise_bad_password()

    pdf.save(output_path)
    pdf.close()

    unlocked_size = os.path.getsize(output_path)
    cleanup_temp(input_path)
    return original_size, unlocked_size


def pdf_to_jpg_service(
    input_path: Path, output_folder: Path, quality: str, page_mode: str, pages: str
) -> Tuple[List[Path], int]:
    try:
        import fitz
    except ImportError:
        raise_internal_error("Dependencia nao instalada: pymupdf")

    dpi_map = {"low": 72, "medium": 150, "high": 300}
    dpi = dpi_map.get(quality, 150)
    zoom = dpi / 72

    doc = fitz.open(str(input_path))
    total_pages = doc.page_count

    if page_mode == "specific" and pages.strip():
        page_indexes = parse_page_range(pages, total_pages)
        page_indexes = [p - 1 for p in page_indexes]
    else:
        page_indexes = list(range(total_pages))

    image_paths = []
    mat = fitz.Matrix(zoom, zoom)
    for idx in page_indexes:
        page = doc[idx]
        pix = page.get_pixmap(matrix=mat, alpha=False)
        img_path = output_folder / f"pagina_{idx + 1:03d}.jpg"
        pix.save(str(img_path))
        image_paths.append(img_path)

    doc.close()

    if not image_paths:
        raise_bad_request("Nenhuma pagina valida para converter")

    return image_paths, len(image_paths)


def crop_pdf_service(
    input_path: Path, output_path: Path,
    x: float, y: float, width: float, height: float, pages: str
) -> Tuple[int, int]:
    if width <= 0 or height <= 0:
        raise_bad_request("Largura e altura devem ser maiores que zero")

    try:
        import fitz
    except ImportError:
        raise_internal_error("Dependencia nao instalada: pymupdf")

    doc = fitz.open(str(input_path))
    total_pages = doc.page_count
    page_list = parse_page_range(pages, total_pages)

    for i in range(total_pages):
        if (i + 1) in page_list:
            page = doc[i]
            page_rect = page.rect

            crop_left = max(0, min(x, page_rect.width))
            crop_top = max(0, min(y, page_rect.height))
            crop_right = max(crop_left + 1, min(x + width, page_rect.width))
            crop_bottom = max(crop_top + 1, min(y + height, page_rect.height))

            crop_rect = fitz.Rect(crop_left, crop_top, crop_right, crop_bottom)
            page.set_cropbox(crop_rect)

    doc.save(output_path, garbage=3, deflate=True)
    doc.close()
    cleanup_temp(input_path)
    return total_pages, len(page_list)


def pdf_info_service(input_path: Path, filename: str, content_size: int) -> dict:
    reader = PdfReader(input_path)
    total_pages = len(reader.pages)
    meta = reader.metadata or {}

    return {
        "filename": filename,
        "total_pages": total_pages,
        "file_size_bytes": content_size,
        "encrypted": reader.is_encrypted,
        "title": str(meta.get('/Title', '')),
        "author": str(meta.get('/Author', '')),
        "subject": str(meta.get('/Subject', '')),
        "creator": str(meta.get('/Creator', '')),
        "producer": str(meta.get('/Producer', '')),
    }


def read_metadata_service(input_path: Path, filename: str, content_size: int) -> dict:
    reader = PdfReader(input_path)
    meta = reader.metadata or {}
    return {
        "title": str(meta.get('/Title', '')),
        "author": str(meta.get('/Author', '')),
        "subject": str(meta.get('/Subject', '')),
        "creator": str(meta.get('/Creator', '')),
        "producer": str(meta.get('/Producer', '')),
        "total_pages": len(reader.pages),
        "file_size_bytes": content_size,
    }


def edit_metadata_service(input_path: Path, output_path: Path, title: str, author: str, subject: str) -> None:
    reader = PdfReader(input_path)
    writer = PdfWriter()

    for page in reader.pages:
        writer.add_page(page)

    metadata = {}
    if title:
        metadata['/Title'] = title
    if author:
        metadata['/Author'] = author
    if subject:
        metadata['/Subject'] = subject

    writer.add_metadata(metadata)

    with open(output_path, "wb") as f:
        writer.write(f)

    cleanup_temp(input_path)


def pdf_to_png_service(
    input_path: Path, output_folder: Path, quality: str, page_mode: str, pages: str
) -> Tuple[List[Path], int]:
    try:
        import fitz
    except ImportError:
        raise_internal_error("Dependencia nao instalada: pymupdf")

    dpi_map = {"low": 72, "medium": 150, "high": 300}
    dpi = dpi_map.get(quality, 150)
    zoom = dpi / 72

    doc = fitz.open(str(input_path))
    total_pages = doc.page_count

    if page_mode == "specific" and pages.strip():
        page_indexes = parse_page_range(pages, total_pages)
        page_indexes = [p - 1 for p in page_indexes]
    else:
        page_indexes = list(range(total_pages))

    image_paths = []
    mat = fitz.Matrix(zoom, zoom)
    for idx in page_indexes:
        page = doc[idx]
        pix = page.get_pixmap(matrix=mat, alpha=False)
        img_path = output_folder / f"pagina_{idx + 1:03d}.png"
        pix.save(str(img_path))
        image_paths.append(img_path)

    doc.close()

    if not image_paths:
        raise_bad_request("Nenhuma pagina valida para converter")

    return image_paths, len(image_paths)


def pdf_to_webp_service(
    input_path: Path, output_folder: Path, quality: str, page_mode: str, pages: str
) -> Tuple[List[Path], int]:
    try:
        import fitz
        from PIL import Image
    except ImportError:
        raise_internal_error("Dependencia nao instalada: pymupdf, pillow")

    dpi_map = {"low": 72, "medium": 150, "high": 300}
    dpi = dpi_map.get(quality, 150)
    zoom = dpi / 72

    doc = fitz.open(str(input_path))
    total_pages = doc.page_count

    if page_mode == "specific" and pages.strip():
        page_indexes = parse_page_range(pages, total_pages)
        page_indexes = [p - 1 for p in page_indexes]
    else:
        page_indexes = list(range(total_pages))

    image_paths = []
    mat = fitz.Matrix(zoom, zoom)
    for idx in page_indexes:
        page = doc[idx]
        pix = page.get_pixmap(matrix=mat, alpha=False)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        img_path = output_folder / f"pagina_{idx + 1:03d}.webp"
        img.save(str(img_path), format="WEBP", quality=85)
        image_paths.append(img_path)

    doc.close()

    if not image_paths:
        raise_bad_request("Nenhuma pagina valida para converter")

    return image_paths, len(image_paths)


def extract_images_service(input_path: Path, output_folder: Path) -> Tuple[List[Path], int]:
    try:
        import fitz
    except ImportError:
        raise_internal_error("Dependencia nao instalada: pymupdf")

    doc = fitz.open(str(input_path))
    image_paths = []
    count = 0

    for page_idx in range(doc.page_count):
        page = doc[page_idx]
        image_list = page.get_images(full=True)
        for img_idx, img in enumerate(image_list):
            xref = img[0]
            base_image = doc.extract_image(xref)
            image_bytes = base_image["image"]
            ext = base_image["ext"]
            img_name = f"pagina_{page_idx + 1:03d}_imagem_{img_idx + 1:03d}.{ext}"
            img_path = output_folder / img_name
            with open(img_path, "wb") as f:
                f.write(image_bytes)
            image_paths.append(img_path)
            count += 1

    doc.close()

    if not image_paths:
        raise_bad_request("Nenhuma imagem encontrada no PDF")

    return image_paths, count


def pdf_to_excel_service(
    input_path: Path, output_path: Path, mode: str, pages: str
) -> Tuple[Path, int]:
    try:
        import pdfplumber
        import pandas as pd
        from openpyxl import Workbook
    except ImportError:
        raise_internal_error("Dependencias nao instaladas: pdfplumber, pandas, openpyxl")

    from pypdf import PdfReader

    reader = PdfReader(str(input_path))
    total_pages = len(reader.pages)

    if pages.strip():
        page_list = parse_page_range(pages, total_pages)
    else:
        page_list = list(range(1, total_pages + 1))

    tables_found = 0

    with pdfplumber.open(str(input_path)) as pdf:
        if mode == 'tables':
            with pd.ExcelWriter(str(output_path), engine='openpyxl') as writer_obj:
                has_data = False
                for page_num in page_list:
                    page = pdf.pages[page_num - 1]
                    tables = page.extract_tables()
                    if tables:
                        for table_idx, table in enumerate(tables):
                            if table and any(any(cell for cell in row if cell) for row in table):
                                headers = table[0] if table else []
                                data = table[1:] if len(table) > 1 else []
                                df = pd.DataFrame(data, columns=headers if headers and all(headers) else None)
                                sheet_name = f"pag{page_num}_tab{table_idx + 1}"[:31]
                                df.to_excel(writer_obj, sheet_name=sheet_name, index=False)
                                tables_found += 1
                                has_data = True
                    else:
                        text = page.extract_text()
                        if text:
                            lines = text.split('\n')
                            df = pd.DataFrame({'linha': range(1, len(lines) + 1), 'texto': lines})
                            sheet_name = f"pag{page_num}_texto"[:31]
                            df.to_excel(writer_obj, sheet_name=sheet_name, index=False)
                            has_data = True

                if not has_data:
                    raise_bad_request("Nenhuma tabela ou texto encontrado neste PDF")
        else:
            rows = []
            for page_num in page_list:
                page = pdf.pages[page_num - 1]
                text = page.extract_text()
                if text:
                    for line_num, line in enumerate(text.split('\n'), 1):
                        rows.append({'pagina': page_num, 'linha': line_num, 'texto': line})

            if not rows:
                raise_bad_request("Nenhum texto encontrado neste PDF")

            df = pd.DataFrame(rows)
            df.to_excel(str(output_path), index=False)

    return output_path, tables_found


def excel_to_pdf_service(input_path: Path, output_path: Path) -> Tuple[Path, int]:
    try:
        import openpyxl
    except ImportError:
        raise_internal_error("Dependencia nao instalada: openpyxl")

    wb = openpyxl.load_workbook(str(input_path), data_only=True)
    sheet_count = len(wb.sheetnames)

    try:
        import fitz
    except ImportError:
        raise_internal_error("Dependencia nao instalada: pymupdf")

    doc = fitz.open()
    margin = 50
    usable_width = 595 - 2 * margin

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        max_rows = len(rows)
        max_cols = max(len(row) for row in rows)
        col_width = max(usable_width // max(max_cols, 1), 40)

        rows_per_page = int((842 - 2 * margin) / 20)
        pages_needed = max(1, (max_rows + rows_per_page - 1) // rows_per_page)

        for page_idx in range(pages_needed):
            page = doc.new_page(width=595, height=842)
            start_row = page_idx * rows_per_page
            end_row = min(start_row + rows_per_page + 1, max_rows)

            y = margin
            x = margin

            if page_idx == 0:
                page.insert_text(fitz.Point(margin, margin - 20),
                                 f"Planilha: {sheet_name}",
                                 fontsize=14, fontname="helv", color=(0, 0, 0))

            table_data = []
            for row_idx, row in enumerate(rows[start_row:end_row]):
                table_data.append([str(cell) if cell is not None else "" for cell in row])

            for row_idx, row_data in enumerate(table_data):
                for col_idx, cell in enumerate(row_data):
                    x_pos = margin + col_idx * col_width
                    if x_pos + col_width > 595 - margin:
                        break
                    rect = fitz.Rect(x_pos, y, x_pos + col_width, y + 18)
                    page.draw_rect(rect, color=(0.8, 0.8, 0.8), width=0.5)
                    if row_idx == 0:
                        page.draw_rect(rect, color=(0.9, 0.9, 0.9), fill=(0.9, 0.9, 0.9))
                    font = "hebo" if row_idx == 0 else "helv"
                    font_size = 10 if row_idx == 0 else 9
                    text = str(cell)[:int(col_width / 5)]
                    page.insert_textbox(rect + (-2, 2, -2, -2),
                                        text, fontsize=font_size,
                                        fontname=font, color=(0, 0, 0))
                y += 18
                if y > 842 - margin:
                    break

    doc.save(str(output_path), garbage=3, deflate=True)
    doc.close()
    wb.close()
    return output_path, sheet_count


def pdf_to_txt_service(input_path: Path, output_path: Path) -> Tuple[Path, int, int]:
    try:
        import fitz
    except ImportError:
        raise_internal_error("Dependencia nao instalada: pymupdf")

    doc = fitz.open(str(input_path))
    page_count = doc.page_count
    chars = 0

    with open(str(output_path), "w", encoding="utf-8") as f:
        for i in range(page_count):
            page = doc[i]
            text = page.get_text()
            if text:
                f.write(text)
                f.write("\n\n")
                chars += len(text)

    doc.close()
    return output_path, page_count, chars
